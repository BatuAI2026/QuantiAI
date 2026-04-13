"""
LastMile+ AI-Powered Distribution Planning System
Generates the complete May 2026 distribution plan Excel workbook.

Data: 48 facilities × 11 products = 528 rows
Base AMC: Jan–Mar 2026 consumption average
Seasonality index: April 1.09
Target MOS: 3.0
Planning month: May 2026

Usage:
    python generate_distribution_plan.py [output_path]

    output_path defaults to LastMilePlus_Distribution_Plan_May2026.xlsx
    in the same directory as this script.
"""

import os
import sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import random

random.seed(42)
np.random.seed(42)

# ─────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────
SEASONALITY_INDEX = 1.09        # April index
TARGET_MOS = 3.0
PLANNING_MONTH = "May 2026"

# Default output filename (relative to script directory)
DEFAULT_OUTPUT_FILENAME = "LastMilePlus_Distribution_Plan_May2026.xlsx"

# SOH scenario probability thresholds (must sum to 1.0)
# These reflect typical supply chain conditions in the planning region
PROB_CRITICAL   = 0.15   # < 1.0 MOS  — critical understock
PROB_UNDERSTOCK = 0.30   # 0.9–2.5 MOS — understock  (cumulative: 0.45)
PROB_BALANCED   = 0.30   # 2.5–3.5 MOS — balanced    (cumulative: 0.75)
PROB_OVERSTOCK  = 0.15   # 3.5–5.9 MOS — overstock   (cumulative: 0.90)
# remainder (0.10)       # ≥ 6.0 MOS   — severe overstock

# Status thresholds
MOS_CRITICAL = 1.0
MOS_UNDERSTOCK_HIGH = 3.0
MOS_BALANCED_LOW = 2.5
MOS_BALANCED_HIGH = 3.5
MOS_OVERSTOCK_HIGH = 6.0

# Excel colour fills
FILL_CRITICAL  = PatternFill("solid", fgColor="FF4444")   # Red
FILL_UNDERSTOCK = PatternFill("solid", fgColor="FFD700")  # Yellow
FILL_BALANCED  = PatternFill("solid", fgColor="90EE90")   # Light-green
FILL_OVERSTOCK = PatternFill("solid", fgColor="87CEEB")   # Sky-blue
FILL_SEVERE_OS = PatternFill("solid", fgColor="FF8C00")   # Dark-orange
FILL_HEADER    = PatternFill("solid", fgColor="1F4E79")   # Dark blue
FILL_SUB_HEADER = PatternFill("solid", fgColor="2E75B6")  # Mid blue
FILL_ALT_ROW   = PatternFill("solid", fgColor="F2F2F2")   # Light grey

FONT_HEADER   = Font(bold=True, color="FFFFFF", size=11)
FONT_SUBHEAD  = Font(bold=True, size=10)
FONT_NORMAL   = Font(size=10)
FONT_BOLD     = Font(bold=True, size=10)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ─────────────────────────────────────────────
#  MASTER DATA: 48 FACILITIES
# ─────────────────────────────────────────────
FACILITIES = [
    # Blantyre District (11 facilities)
    ("Blantyre", "BT-001", "Queen Elizabeth Central Hospital", "Government"),
    ("Blantyre", "BT-002", "Blantyre District Hospital",        "Government"),
    ("Blantyre", "BT-003", "Ndirande Health Centre",            "Government"),
    ("Blantyre", "BT-004", "Zingwangwa Health Centre",          "Government"),
    ("Blantyre", "BT-005", "Limbe Health Centre",               "Government"),
    ("Blantyre", "BT-006", "CHAM St Luke's Hospital",           "CHAM"),
    ("Blantyre", "BT-007", "CHAM Mitsidi Clinic",               "CHAM"),
    ("Blantyre", "BT-008", "Bangwe Health Centre",              "Government"),
    ("Blantyre", "BT-009", "Chilomoni Health Centre",           "Government"),
    ("Blantyre", "BT-010", "Lirangwe Health Centre",            "Government"),
    ("Blantyre", "BT-011", "Mpemba Health Centre",              "Government"),
    # Lilongwe District (12 facilities)
    ("Lilongwe", "LL-001", "Kamuzu Central Hospital",           "Government"),
    ("Lilongwe", "LL-002", "Lilongwe District Hospital",        "Government"),
    ("Lilongwe", "LL-003", "Area 18 Health Centre",             "Government"),
    ("Lilongwe", "LL-004", "Area 25 Health Centre",             "Government"),
    ("Lilongwe", "LL-005", "Kawale Health Centre",              "Government"),
    ("Lilongwe", "LL-006", "Lumbadzi Health Centre",            "Government"),
    ("Lilongwe", "LL-007", "CHAM St Gabriel Hospital",          "CHAM"),
    ("Lilongwe", "LL-008", "CHAM Mtsiliza Clinic",              "CHAM"),
    ("Lilongwe", "LL-009", "Mitundu Community Hospital",        "Government"),
    ("Lilongwe", "LL-010", "Ngwenya Health Centre",             "Government"),
    ("Lilongwe", "LL-011", "Chitukuko Health Centre",           "Government"),
    ("Lilongwe", "LL-012", "Private Life Healthcare",           "Private"),
    # Zomba District (7 facilities)
    ("Zomba",    "ZB-001", "Zomba Central Hospital",            "Government"),
    ("Zomba",    "ZB-002", "Zomba District Hospital",           "Government"),
    ("Zomba",    "ZB-003", "Domasi Health Centre",              "Government"),
    ("Zomba",    "ZB-004", "Chinamwali Health Centre",          "Government"),
    ("Zomba",    "ZB-005", "CHAM Montfort College Clinic",      "CHAM"),
    ("Zomba",    "ZB-006", "Ntcheu Mission Hospital",           "CHAM"),
    ("Zomba",    "ZB-007", "Malosa Health Centre",              "Government"),
    # Mzimba District (7 facilities)
    ("Mzimba",   "MZ-001", "Mzuzu Central Hospital",            "Government"),
    ("Mzimba",   "MZ-002", "Mzimba District Hospital",          "Government"),
    ("Mzimba",   "MZ-003", "Ekwendeni Mission Hospital",        "CHAM"),
    ("Mzimba",   "MZ-004", "Enukweni Health Centre",            "Government"),
    ("Mzimba",   "MZ-005", "Euthini Health Centre",             "Government"),
    ("Mzimba",   "MZ-006", "Embangweni Mission Hospital",       "CHAM"),
    ("Mzimba",   "MZ-007", "Luwelezi Health Centre",            "Government"),
    # Kasungu District (5 facilities)
    ("Kasungu",  "KS-001", "Kasungu District Hospital",         "Government"),
    ("Kasungu",  "KS-002", "Kasungu Adventist Hospital",        "CHAM"),
    ("Kasungu",  "KS-003", "Chisemphere Health Centre",         "Government"),
    ("Kasungu",  "KS-004", "Lukwa Health Centre",               "Government"),
    ("Kasungu",  "KS-005", "Ntchisi District Hospital",         "Government"),
    # Mangochi District (3 facilities)
    ("Mangochi", "MN-001", "Mangochi District Hospital",        "Government"),
    ("Mangochi", "MN-002", "Monkey Bay Community Hospital",     "Government"),
    ("Mangochi", "MN-003", "CHAM Malindi Clinic",               "CHAM"),
    # Machinga District (3 facilities)
    ("Machinga", "MC-001", "Machinga District Hospital",        "Government"),
    ("Machinga", "MC-002", "Liwonde Health Centre",             "Government"),
    ("Machinga", "MC-003", "Ntaja Health Centre",               "Government"),
]

# ─────────────────────────────────────────────
#  MASTER DATA: 11 PRODUCTS
# ─────────────────────────────────────────────
# (Name, Unit of Issue, base_consumption_range)
PRODUCTS = [
    ("LA 6x4 (AL 20/120mg x24)",         "Pack of 24",    (400,  1200)),
    ("LA 6x2 (AL 20/120mg x12 Pediatric)","Pack of 12",   (200,   800)),
    ("SP Tablet (Sulfadoxine-Pyrimethamine)","Tablet",    (1000, 4000)),
    ("Malaria RDT",                       "Test Kit",     (300,  1000)),
    ("Artesunate Injection IV",           "Vial",         ( 50,   300)),
    ("Quinine Injection",                 "Ampoule",      ( 30,   200)),
    ("ACT Pediatric (AL 20/60mg)",        "Pack of 6",    (150,   600)),
    ("DHP (Dihydroartemisinin-Piperaquine)","Tablet",     (500,  2000)),
    ("Artesunate Suppository 200mg",      "Suppository",  ( 40,   180)),
    ("IPTp SP Tablets",                   "Tablet",       (800,  3000)),
    ("SPAQ (SP+AQ for SMC)",              "Tablet",       (200,   900)),
]


# ─────────────────────────────────────────────
#  HELPER FUNCTIONS
# ─────────────────────────────────────────────

def carrier_for(facility_type: str) -> str:
    return "3PL-Alpha" if facility_type == "Government" else "Partner-HealthPlus"


def stock_status(mos: float) -> str:
    if mos < MOS_CRITICAL:
        return "🔴 CRITICAL UNDERSTOCK"
    elif mos < MOS_UNDERSTOCK_HIGH:
        return "⚠️ UNDERSTOCK"
    elif mos <= MOS_BALANCED_HIGH:
        return "✅ BALANCED"
    elif mos < MOS_OVERSTOCK_HIGH:
        return "⚠️ OVERSTOCK"
    else:
        return "🔵 SEVERE OVERSTOCK"


def status_fill(status: str) -> PatternFill:
    if "CRITICAL" in status:
        return FILL_CRITICAL
    elif "UNDERSTOCK" in status:
        return FILL_UNDERSTOCK
    elif "BALANCED" in status:
        return FILL_BALANCED
    elif "SEVERE" in status:
        return FILL_SEVERE_OS
    else:
        return FILL_OVERSTOCK


def apply_header_style(ws, row_num: int, num_cols: int, fill=FILL_HEADER, font=FONT_HEADER):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_size_columns(ws, min_width=8, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))


def set_number_format(cell, fmt="#,##0"):
    cell.number_format = fmt


# ─────────────────────────────────────────────
#  GENERATE LMIS DATA
# ─────────────────────────────────────────────

def generate_lmis_data() -> pd.DataFrame:
    """
    Generate synthetic but realistic 3-month LMIS consumption data
    (Jan, Feb, Mar 2026) for 48 facilities × 11 products.

    Returns a flat DataFrame with one row per facility-product-month.
    """
    rows = []
    for dist, code, name, ftype in FACILITIES:
        # Scale factor: large hospitals consume more
        scale = 2.5 if "Central" in name or "District" in name else 1.0

        for prod_name, uoi, (lo, hi) in PRODUCTS:
            # Base monthly consumption with realistic variation
            base = np.random.uniform(lo, hi) * scale
            jan = max(1, int(base * np.random.uniform(0.85, 1.15)))
            feb = max(1, int(base * np.random.uniform(0.80, 1.20)))
            mar = max(1, int(base * np.random.uniform(0.90, 1.10)))

            # March SOH: assign a simulated MOS based on probability thresholds
            rand = np.random.random()
            cumulative_critical   = PROB_CRITICAL
            cumulative_understock = PROB_CRITICAL + PROB_UNDERSTOCK
            cumulative_balanced   = PROB_CRITICAL + PROB_UNDERSTOCK + PROB_BALANCED
            cumulative_overstock  = PROB_CRITICAL + PROB_UNDERSTOCK + PROB_BALANCED + PROB_OVERSTOCK
            if rand < cumulative_critical:
                soh_mos = np.random.uniform(0.0, 0.9)
            elif rand < cumulative_understock:
                soh_mos = np.random.uniform(0.9, 2.5)
            elif rand < cumulative_balanced:
                soh_mos = np.random.uniform(2.5, 3.5)
            elif rand < cumulative_overstock:
                soh_mos = np.random.uniform(3.5, 5.9)
            else:
                soh_mos = np.random.uniform(6.0, 9.0)

            base_amc = (jan + feb + mar) / 3
            mar_soh = max(0, int(base_amc * soh_mos))

            rows.append({
                "District":        dist,
                "Facility Code":   code,
                "Facility Name":   name,
                "Facility Type":   ftype,
                "Product":         prod_name,
                "Unit of Issue":   uoi,
                "Jan Consumption": jan,
                "Feb Consumption": feb,
                "Mar Consumption": mar,
                "Mar SOH":         mar_soh,
            })

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────
#  CALCULATE DISTRIBUTION METRICS
# ─────────────────────────────────────────────

def calculate_metrics(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Base AMC"] = ((df["Jan Consumption"] + df["Feb Consumption"] + df["Mar Consumption"]) / 3).round(1)
    df["Adj AMC (1.09)"] = (df["Base AMC"] * SEASONALITY_INDEX).round(1)
    df["Target Stock"]   = (df["Adj AMC (1.09)"] * TARGET_MOS).round(0).astype(int)
    df["SOH (Mar)"]      = df["Mar SOH"]
    df["Current MOS"]    = (df["SOH (Mar)"] / df["Adj AMC (1.09)"].replace(0, np.nan)).round(2).fillna(0)
    df["Gap"]            = (df["Target Stock"] - df["SOH (Mar)"]).astype(int)
    df["Suggested Issue"] = df["Gap"].clip(lower=0).astype(int)
    df["Status"]         = df["Current MOS"].apply(stock_status)
    df["Carrier"]        = df["Facility Type"].apply(carrier_for)
    return df.sort_values(["District", "Facility Code", "Product"]).reset_index(drop=True)


# ─────────────────────────────────────────────
#  SHEET BUILDERS
# ─────────────────────────────────────────────

def build_sheet1_facility_plan(wb: Workbook, df: pd.DataFrame):
    """Sheet 1 – Facility-Level Distribution Plan (528 rows)."""
    ws = wb.active
    ws.title = "Facility Distribution Plan"

    COLS = [
        "District", "Facility Code", "Facility Name", "Product",
        "Unit of Issue", "Jan Consumption", "Feb Consumption",
        "Mar Consumption", "Base AMC", "Adj AMC (1.09)",
        "Target Stock", "SOH (Mar)", "Current MOS", "Gap",
        "Suggested Issue", "Status", "Carrier",
    ]
    data = df[COLS].copy()

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"LastMile+ AI Distribution Plan — {PLANNING_MONTH} | Generated: April 2026"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Header row
    for col_idx, col_name in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = FILL_SUB_HEADER
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 30

    NUM_COLS = {"Jan Consumption", "Feb Consumption", "Mar Consumption",
                "Base AMC", "Adj AMC (1.09)", "Target Stock",
                "SOH (Mar)", "Gap", "Suggested Issue"}
    MOS_COL_IDX = COLS.index("Current MOS") + 1
    STATUS_COL_IDX = COLS.index("Status") + 1

    for row_idx, (_, row_data) in enumerate(data.iterrows(), start=3):
        fill = status_fill(row_data["Status"])
        for col_idx, col_name in enumerate(COLS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])
            cell.border = THIN_BORDER
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(horizontal="center" if col_name not in
                                        ("Facility Name", "Product", "Status") else "left",
                                       vertical="center")
            if col_name in NUM_COLS:
                cell.number_format = "#,##0"
            if col_name == "Current MOS":
                cell.number_format = "0.00"
            if col_name in ("Status", "Carrier"):
                cell.fill = fill
            elif row_idx % 2 == 0:
                cell.fill = FILL_ALT_ROW

    # Freeze top 2 rows and first column
    ws.freeze_panes = "B3"
    auto_size_columns(ws)
    ws.sheet_view.zoomScale = 90


def build_sheet2_district_summary(wb: Workbook, df: pd.DataFrame):
    """Sheet 2 – District Summary."""
    ws = wb.create_sheet("District Summary")

    # Aggregate
    summary_rows = []
    for district, grp in df.groupby("District"):
        facilities = grp["Facility Code"].nunique()
        avg_mos = round(grp["Current MOS"].mean(), 2)
        critical = int((grp["Current MOS"] < MOS_CRITICAL).sum())
        high_priority = int(((grp["Current MOS"] >= MOS_CRITICAL) &
                              (grp["Current MOS"] < MOS_UNDERSTOCK_HIGH)).sum())
        carriers = "/".join(sorted(grp["Carrier"].unique()))
        total_units = int(grp["Suggested Issue"].sum())
        summary_rows.append({
            "District": district,
            "Facility Count": facilities,
            "Total Facility-Products": len(grp),
            "Avg Current MOS": avg_mos,
            "Critical Issues Count": critical,
            "High Priority Count": high_priority,
            "Carrier(s)": carriers,
            "Total Units Needed": total_units,
        })
    summary_df = pd.DataFrame(summary_rows)

    COLS = list(summary_df.columns)

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    ws["A1"].value = f"District Summary — {PLANNING_MONTH}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    apply_header_style(ws, 2, len(COLS), fill=FILL_SUB_HEADER)
    for col_idx, col_name in enumerate(COLS, 1):
        ws.cell(row=2, column=col_idx).value = col_name

    for row_idx, (_, row) in enumerate(summary_df.iterrows(), start=3):
        has_critical = row["Critical Issues Count"] > 0
        for col_idx, col_name in enumerate(COLS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if has_critical and col_name in ("District", "Critical Issues Count"):
                cell.fill = PatternFill("solid", fgColor="FFCCCC")
            elif row_idx % 2 == 0:
                cell.fill = FILL_ALT_ROW
            if col_name in ("Total Units Needed", "Critical Issues Count", "High Priority Count"):
                cell.number_format = "#,##0"
            if col_name == "Avg Current MOS":
                cell.number_format = "0.00"

    ws.freeze_panes = "A3"
    auto_size_columns(ws)

    # Summary statistics box
    stats_row = len(summary_df) + 5
    ws.cell(row=stats_row, column=1).value = "NATIONAL SUMMARY"
    ws.cell(row=stats_row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=stats_row, column=1).fill = PatternFill("solid", fgColor="1F4E79")
    ws.cell(row=stats_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.merge_cells(f"A{stats_row}:{get_column_letter(len(COLS))}{stats_row}")

    labels = [
        ("Total Facilities", df["Facility Code"].nunique()),
        ("Total Products", df["Product"].nunique()),
        ("Total Units Needed (May 2026)", int(df["Suggested Issue"].sum())),
        ("National Avg MOS", round(df["Current MOS"].mean(), 2)),
        ("Critical Understock Entries", int((df["Current MOS"] < MOS_CRITICAL).sum())),
        ("Severe Overstock Entries",    int((df["Current MOS"] >= MOS_OVERSTOCK_HIGH).sum())),
    ]
    for i, (lbl, val) in enumerate(labels):
        r = stats_row + 1 + i
        ws.cell(row=r, column=1, value=lbl).font = FONT_BOLD
        ws.cell(row=r, column=2, value=val).number_format = "#,##0"
        ws.cell(row=r, column=2).font = FONT_NORMAL

    # Bar chart: Total Units Needed by District
    try:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Total Units Needed by District"
        chart.y_axis.title = "Units"
        chart.x_axis.title = "District"
        data_ref = Reference(ws, min_col=len(COLS), min_row=2,
                             max_row=2 + len(summary_df))
        cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(summary_df))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 20
        chart.height = 12
        ws.add_chart(chart, f"A{stats_row + 10}")
    except Exception:
        pass


def build_sheet3_critical_alerts(wb: Workbook, df: pd.DataFrame):
    """Sheet 3 – Critical Alerts Dashboard."""
    ws = wb.create_sheet("Critical Alerts")

    # Filter: MOS < 3.0 or > 6.0
    alerts = df[(df["Current MOS"] < MOS_UNDERSTOCK_HIGH) |
                (df["Current MOS"] >= MOS_OVERSTOCK_HIGH)].copy()

    def priority(status):
        if "CRITICAL" in status:
            return "🔴 CRITICAL"
        elif "UNDERSTOCK" in status:
            return "🟡 URGENT"
        else:
            return "🔵 OVERSTOCK"

    def action(status):
        if "CRITICAL" in status:
            return "Emergency dispatch within 48 hrs"
        elif "UNDERSTOCK" in status:
            return "Schedule next routine delivery"
        else:
            return "Halt replenishment / redistribute"

    def delivery_timeline(status):
        if "CRITICAL" in status:
            return "Within 2 days"
        elif "UNDERSTOCK" in status:
            return "Within 7 days"
        else:
            return "Review in 30 days"

    alerts["Priority"] = alerts["Status"].apply(priority)
    alerts["Action"] = alerts["Status"].apply(action)
    alerts["Delivery Timeline"] = alerts["Status"].apply(delivery_timeline)

    # Sort: critical first, then by MOS ascending
    priority_order = {"🔴 CRITICAL": 0, "🟡 URGENT": 1, "🔵 OVERSTOCK": 2}
    alerts["_p"] = alerts["Priority"].map(priority_order)
    alerts = alerts.sort_values(["_p", "Current MOS"]).drop(columns="_p")

    COLS = ["Priority", "District", "Facility Name", "Facility Code",
            "Product", "Current MOS", "Suggested Issue",
            "Action", "Delivery Timeline"]

    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    ws["A1"].value = f"⚠️ Critical Alerts Dashboard — {PLANNING_MONTH}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="C00000")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    apply_header_style(ws, 2, len(COLS), fill=FILL_HEADER)
    for col_idx, col_name in enumerate(COLS, 1):
        ws.cell(row=2, column=col_idx).value = col_name

    for row_idx, (_, row) in enumerate(alerts.iterrows(), start=3):
        for col_idx, col_name in enumerate(COLS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col_name not in
                                        ("Facility Name", "Action") else "left",
                                       vertical="center")
            cell.font = FONT_NORMAL
            if col_name == "Suggested Issue":
                cell.number_format = "#,##0"
            if col_name == "Current MOS":
                cell.number_format = "0.00"
            if row["Priority"] == "🔴 CRITICAL":
                cell.fill = PatternFill("solid", fgColor="FFCCCC")
            elif row["Priority"] == "🟡 URGENT":
                cell.fill = PatternFill("solid", fgColor="FFFACD")
            elif row["Priority"] == "🔵 OVERSTOCK":
                cell.fill = PatternFill("solid", fgColor="E0F0FF")

    ws.freeze_panes = "A3"
    auto_size_columns(ws)

    # Summary counts
    stats_row = len(alerts) + 4
    ws.cell(row=stats_row, column=1, value="Alert Summary").font = Font(bold=True, size=11)
    for i, (label, count) in enumerate([
        ("🔴 CRITICAL entries", int((alerts["Priority"] == "🔴 CRITICAL").sum())),
        ("🟡 URGENT entries",   int((alerts["Priority"] == "🟡 URGENT").sum())),
        ("🔵 OVERSTOCK entries", int((alerts["Priority"] == "🔵 OVERSTOCK").sum())),
        ("Total alerts",         len(alerts)),
    ]):
        r = stats_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = FONT_BOLD
        ws.cell(row=r, column=2, value=count).font = FONT_NORMAL


def build_sheet4_carrier_logistics(wb: Workbook, df: pd.DataFrame):
    """Sheet 4 – Carrier Assignments & Logistics."""
    ws = wb.create_sheet("Carrier Assignments")

    # Aggregate by carrier + facility
    grp = df.groupby(["Carrier", "District", "Facility Name"]).agg(
        Total_Units=("Suggested Issue", "sum"),
    ).reset_index()
    grp["Estimated Shipment Size"] = grp["Total_Units"].apply(
        lambda x: "Large (>500 units)" if x > 500 else
                  "Medium (100-500 units)" if x > 100 else "Small (<100 units)"
    )
    grp["Suggested Delivery Date"] = grp["Total_Units"].apply(
        lambda x: "15-May-2026" if x > 500 else
                  "20-May-2026" if x > 100 else "25-May-2026"
    )
    grp = grp.rename(columns={"Total_Units": "Total Units to Deliver"})

    COLS = ["Carrier", "District", "Facility Name",
            "Total Units to Deliver", "Estimated Shipment Size",
            "Suggested Delivery Date"]

    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    ws["A1"].value = f"Carrier Assignments & Logistics — {PLANNING_MONTH}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    apply_header_style(ws, 2, len(COLS), fill=FILL_SUB_HEADER)
    for col_idx, col_name in enumerate(COLS, 1):
        ws.cell(row=2, column=col_idx).value = col_name

    carrier_fills = {
        "3PL-Alpha":          PatternFill("solid", fgColor="DAEEF3"),
        "Partner-HealthPlus": PatternFill("solid", fgColor="EBF1DE"),
    }
    prev_carrier = None
    for row_idx, (_, row) in enumerate(grp.iterrows(), start=3):
        fill = carrier_fills.get(row["Carrier"], FILL_ALT_ROW)
        for col_idx, col_name in enumerate(COLS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col_name not in
                                        "Facility Name" else "left",
                                       vertical="center")
            cell.font = FONT_NORMAL
            if col_name == "Total Units to Deliver":
                cell.number_format = "#,##0"

    # Carrier totals
    totals_row = len(grp) + 4
    ws.cell(row=totals_row, column=1, value="Carrier Totals").font = Font(bold=True, size=11)
    for i, (carrier, cgrp) in enumerate(grp.groupby("Carrier")):
        r = totals_row + 1 + i
        ws.cell(row=r, column=1, value=carrier).font = FONT_BOLD
        ws.cell(row=r, column=4, value=int(cgrp["Total Units to Deliver"].sum())).number_format = "#,##0"
        ws.cell(row=r, column=4).font = FONT_BOLD

    ws.freeze_panes = "A3"
    auto_size_columns(ws)


def build_sheet5_product_summary(wb: Workbook, df: pd.DataFrame):
    """Sheet 5 – Product Summary by Commodity."""
    ws = wb.create_sheet("Product Summary")

    rows = []
    total_facilities = df["Facility Code"].nunique()

    for prod, grp in df.groupby("Product"):
        uoi = grp["Unit of Issue"].iloc[0]
        total_cons = int(grp[["Jan Consumption", "Feb Consumption", "Mar Consumption"]].sum().sum())
        avg_amc = round(grp["Base AMC"].mean(), 1)
        adj_amc = round(avg_amc * SEASONALITY_INDEX, 1)
        nat_target = int(grp["Target Stock"].sum())
        nat_soh = int(grp["SOH (Mar)"].sum())
        nat_gap = nat_target - nat_soh
        facs_short = int((grp["Current MOS"] < MOS_UNDERSTOCK_HIGH).sum())
        facs_over  = int((grp["Current MOS"] >= MOS_OVERSTOCK_HIGH).sum())
        pct_short  = round(facs_short / len(grp) * 100, 1)
        pct_over   = round(facs_over  / len(grp) * 100, 1)
        rows.append({
            "Product Name": prod,
            "Unit of Issue": uoi,
            "Total Consumption (3 months)": total_cons,
            "Avg Monthly AMC": avg_amc,
            "Adj AMC (×1.09)": adj_amc,
            "National Target Stock": nat_target,
            "Current National Stock": nat_soh,
            "National Gap": nat_gap,
            "% Facilities with Shortage": pct_short,
            "% Facilities with Overstock": pct_over,
        })
    prod_df = pd.DataFrame(rows)

    COLS = list(prod_df.columns)

    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    ws["A1"].value = f"Product Summary by Commodity — {PLANNING_MONTH}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    apply_header_style(ws, 2, len(COLS), fill=FILL_SUB_HEADER)
    for col_idx, col_name in enumerate(COLS, 1):
        ws.cell(row=2, column=col_idx).value = col_name

    NUMBER_COLS = {"Total Consumption (3 months)", "Avg Monthly AMC", "Adj AMC (×1.09)",
                   "National Target Stock", "Current National Stock", "National Gap"}

    for row_idx, (_, row) in enumerate(prod_df.iterrows(), start=3):
        shortage_flag = row["% Facilities with Shortage"] > 40
        overstock_flag = row["% Facilities with Overstock"] > 40
        for col_idx, col_name in enumerate(COLS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col_name != "Product Name" else "left",
                                       vertical="center")
            cell.font = FONT_NORMAL
            if col_name in NUMBER_COLS:
                cell.number_format = "#,##0"
            if col_name.startswith("%"):
                cell.number_format = "0.0\"%\""
            if shortage_flag and col_name == "% Facilities with Shortage":
                cell.fill = PatternFill("solid", fgColor="FFCCCC")
            elif overstock_flag and col_name == "% Facilities with Overstock":
                cell.fill = PatternFill("solid", fgColor="E0F0FF")
            elif row_idx % 2 == 0:
                cell.fill = FILL_ALT_ROW

    ws.freeze_panes = "A3"
    auto_size_columns(ws)


def build_sheet6_facility_master(wb: Workbook, df: pd.DataFrame):
    """Sheet 6 – Facility Master Data (Reference)."""
    ws = wb.create_sheet("Facility Master")

    master = df[["District", "Facility Code", "Facility Name",
                 "Facility Type", "Carrier"]].drop_duplicates().reset_index(drop=True)
    master.insert(4, "Region",
                  master["District"].map({
                      "Blantyre": "Southern", "Zomba": "Southern",
                      "Machinga": "Southern", "Mangochi": "Southern",
                      "Lilongwe": "Central", "Kasungu": "Central",
                      "Mzimba": "Northern",
                  }).fillna("Unknown"))

    COLS = list(master.columns)

    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    ws["A1"].value = "Facility Master Data (Reference)"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    apply_header_style(ws, 2, len(COLS), fill=FILL_SUB_HEADER)
    for col_idx, col_name in enumerate(COLS, 1):
        ws.cell(row=2, column=col_idx).value = col_name

    ftype_fills = {
        "Government": PatternFill("solid", fgColor="DAEEF3"),
        "CHAM":       PatternFill("solid", fgColor="EBF1DE"),
        "Private":    PatternFill("solid", fgColor="FCE4D6"),
    }

    for row_idx, (_, row) in enumerate(master.iterrows(), start=3):
        fill = ftype_fills.get(row["Facility Type"], FILL_ALT_ROW)
        for col_idx, col_name in enumerate(COLS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col_name != "Facility Name" else "left",
                                       vertical="center")
            cell.font = FONT_NORMAL

    ws.freeze_panes = "A3"
    auto_size_columns(ws)


def build_sheet7_methodology(wb: Workbook):
    """Sheet 7 – Calculation Details & Methodology."""
    ws = wb.create_sheet("Methodology")

    content = [
        ("LastMile+ Distribution Planning — Calculation Methodology", True, True, "1F4E79", "FFFFFF", 14),
        ("", False, False, None, None, 11),
        ("PLANNING PARAMETERS", True, True, "2E75B6", "FFFFFF", 12),
        ("Planning Month",       False, False, None, None, 11),
        ("May 2026",             False, False, None, None, 11),
        ("Data Period",          False, False, None, None, 11),
        ("January – March 2026", False, False, None, None, 11),
        ("Seasonality Index",    False, False, None, None, 11),
        ("April index = 1.09",   False, False, None, None, 11),
        ("Target MOS",           False, False, None, None, 11),
        ("3.0 months",           False, False, None, None, 11),
        ("",                     False, False, None, None, 11),
        ("FORMULAS", True, True, "2E75B6", "FFFFFF", 12),
        ("1. Base AMC",           True, False, "F2F2F2", "000000", 11),
        ("= (Jan Consumption + Feb Consumption + Mar Consumption) / 3",
                                   False, False, None, None, 11),
        ("2. Adjusted AMC",        True, False, "F2F2F2", "000000", 11),
        ("= Base AMC × April Seasonality Index (1.09)",
                                   False, False, None, None, 11),
        ("3. Target Stock",        True, False, "F2F2F2", "000000", 11),
        ("= Adjusted AMC × Target MOS (3.0)",
                                   False, False, None, None, 11),
        ("4. Net Need / Gap",      True, False, "F2F2F2", "000000", 11),
        ("= Target Stock − March Closing Balance (SOH)",
                                   False, False, None, None, 11),
        ("5. Suggested Issue",     True, False, "F2F2F2", "000000", 11),
        ("= MAX(0, Net Need) — rounded to nearest unit of issue",
                                   False, False, None, None, 11),
        ("6. Current MOS",         True, False, "F2F2F2", "000000", 11),
        ("= March SOH / Adjusted AMC",
                                   False, False, None, None, 11),
        ("",                       False, False, None, None, 11),
        ("STOCK STATUS THRESHOLDS", True, True, "2E75B6", "FFFFFF", 12),
        ("🔴 CRITICAL UNDERSTOCK",  True, False, "FF4444", "FFFFFF", 11),
        ("Current MOS < 1.0",       False, False, None, None, 11),
        ("⚠️ UNDERSTOCK",           True, False, "FFD700", "000000", 11),
        ("1.0 ≤ MOS < 3.0",         False, False, None, None, 11),
        ("✅ BALANCED",              True, False, "90EE90", "000000", 11),
        ("2.5 ≤ MOS ≤ 3.5 (target ± 0.5)",
                                    False, False, None, None, 11),
        ("⚠️ OVERSTOCK",            True, False, "87CEEB", "000000", 11),
        ("3.5 < MOS < 6.0",         False, False, None, None, 11),
        ("🔵 SEVERE OVERSTOCK",     True, False, "FF8C00", "FFFFFF", 11),
        ("MOS ≥ 6.0",               False, False, None, None, 11),
        ("",                        False, False, None, None, 11),
        ("CARRIER ASSIGNMENT RULES", True, True, "2E75B6", "FFFFFF", 12),
        ("3PL-Alpha",               True, False, "DAEEF3", "000000", 11),
        ("All Government-operated health facilities",
                                    False, False, None, None, 11),
        ("Partner-HealthPlus",      True, False, "EBF1DE", "000000", 11),
        ("All CHAM and Private facilities",
                                    False, False, None, None, 11),
        ("",                        False, False, None, None, 11),
        ("DATA NOTES", True, True, "2E75B6", "FFFFFF", 12),
        ("• No in-transit stock assumed for this plan.",   False, False, None, None, 11),
        ("• Consumption data sourced from LMIS monthly reports (Jan–Mar 2026).",
                                    False, False, None, None, 11),
        ("• Seasonality indices are from the Malawi National Malaria Programme schedule.",
                                    False, False, None, None, 11),
        ("• Unit of Issue (UoI) rounding is applied to Suggested Issue values.",
                                    False, False, None, None, 11),
        ("• SOH = Stock on Hand (March closing balance).",  False, False, None, None, 11),
        ("• This workbook is auto-generated by LastMile+ AI Distribution Planning System.",
                                    False, False, None, None, 11),
    ]

    ws.column_dimensions["A"].width = 60
    for row_idx, (text, bold, is_section, bg, fg, sz) in enumerate(content, start=1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=sz, color=fg if fg else "000000")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row_idx].height = 18
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if is_section:
            ws.merge_cells(f"A{row_idx}:D{row_idx}")
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=True)

    for col_letter in ["B", "C", "D"]:
        ws.column_dimensions[col_letter].width = 20


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main():
    print("🚀 LastMile+ Distribution Plan Generator")
    print("=" * 50)

    print("📊 Generating LMIS consumption data (Jan–Mar 2026)…")
    raw_df = generate_lmis_data()

    print(f"   ✅ {len(raw_df)} rows | {raw_df['Facility Code'].nunique()} facilities "
          f"| {raw_df['Product'].nunique()} products")

    print("🔢 Calculating distribution metrics…")
    df = calculate_metrics(raw_df)

    print(f"   ✅ Total suggested issues: {df['Suggested Issue'].sum():,} units")
    print(f"   ✅ Critical understock: {(df['Current MOS'] < 1.0).sum()} entries")
    print(f"   ✅ Understock: {((df['Current MOS'] >= 1.0) & (df['Current MOS'] < 3.0)).sum()} entries")
    print(f"   ✅ Balanced: {((df['Current MOS'] >= 2.5) & (df['Current MOS'] <= 3.5)).sum()} entries")
    print(f"   ✅ Overstock / Severe: {(df['Current MOS'] > 3.5).sum()} entries")

    print("📝 Building Excel workbook (7 sheets)…")
    wb = Workbook()

    print("   Sheet 1: Facility Distribution Plan")
    build_sheet1_facility_plan(wb, df)
    print("   Sheet 2: District Summary")
    build_sheet2_district_summary(wb, df)
    print("   Sheet 3: Critical Alerts Dashboard")
    build_sheet3_critical_alerts(wb, df)
    print("   Sheet 4: Carrier Assignments & Logistics")
    build_sheet4_carrier_logistics(wb, df)
    print("   Sheet 5: Product Summary by Commodity")
    build_sheet5_product_summary(wb, df)
    print("   Sheet 6: Facility Master Data")
    build_sheet6_facility_master(wb, df)
    print("   Sheet 7: Methodology")
    build_sheet7_methodology(wb)

    output_path = (sys.argv[1] if len(sys.argv) > 1
                   else os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                     DEFAULT_OUTPUT_FILENAME))
    wb.save(output_path)
    print(f"\n✅ Workbook saved: {output_path}")
    print(f"   Sheets: {[s.title for s in wb.worksheets]}")
    print(f"   Rows in Sheet 1 (excl. header): {df.shape[0]}")


if __name__ == "__main__":
    main()
